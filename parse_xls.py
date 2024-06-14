#1. Entferne non DEV/CPP Services
#2. Lösche Übersicht Seite
#3. Speichern als xls
#4. Konvertieren xls to xlsx
#5. Speichern als \Developers\b-end

import openpyxl, win32com.client, calendar, pandas as pd, datetime as dt
from datetime import datetime
from datetime import date

xlapp = win32com.client.DispatchEx("Excel.Application")
print('Aktualisierung xlsx...')
wbr = xlapp.Workbooks.Open('C:\Documents\Statistik Loader.xlsx')
wbr.RefreshAll()
xlapp.CalculateUntilAsyncQueriesDone()
wbr.Save()
print('Aktualisierung ist fertig.')
xlapp.Quit()

print('Lesen die xlsx Datei...')

df = pd.read_excel('Statistik Loader.xlsx', 'Quelldaten', index_col=None, na_values=['NA'])

df = df.drop(df[(df['State'] == 'closed')].index)
df = df.drop(df[(df['Queue'] != '[1] Second Level::DEV') & (df['Queue'] != '[2] Third Level::DEV')].index)
print('Reinigung xlsx...')

writer = pd.ExcelWriter('Statistik.xlsx')
df.to_excel(writer, 'Quelldaten', index=False)
writer.save()

print('Reinigung ist fertig.')

Woerterbuch = { 
				'CPP': 0,
				'JIT': 0,
				'JAVA-Entwicklung': 0,
				'Python': 0,
				'Logistik': 0,
				'EDI-Mapping': 0,
				'EDI': 0,
				'B1-2': 0,
				'System': 0,
				'TP': 0  }
			
Dienste = { '[T000] DEV::JIT': 0 }			

for Wort in Woerterbuch:
	Wort = '[T000] DEV::' + Wort
	Dienste[Wort] = 0

class switch(object):
    value = None
    def __new__(class_, value):
        class_.value = value
        return True

def case(*args):
    return any((arg == switch.value for arg in args))

wb = openpyxl.load_workbook('Statistik.xlsx')

ws = wb['Quelldaten']

print('Kalkulation...')
  
for cell in ws['I']:
	if str(cell.value) in Dienste:
		Dienste [str(cell.value)] += 1

ws = wb.create_sheet('T000')

i = 0
for dienstName in Dienste:
	i += 1
	ws.cell(row=i, column=1).value = dienstName
	ws.cell(row=i, column=2).value = Dienste[dienstName]

currentdate = datetime.today().strftime('%d.%m.%Y')
wb.save('\\0x2srv\DFS-Abteilungen$\Developers\b-end\' + currentdate + '.xlsx')
print('File \\0x2srv\DFS-Abteilungen$\Developers\b-end\' + currentdate + '.xlsx ist gespeichert.')
d,m,y = currentdate.split('.')

woche = dt.date(int(y),int(m),int(d)).strftime("%V")
monat = calendar.month_name[int(m)]

outlook = win32com.client.Dispatch('outlook.application')
mail = outlook.CreateItem(0)
mail.To = 'ruslan@0x22222222.com'
mail.Subject = 'Die Liste Woche #' + woche + ', ' + monat
mail.HtmlBody = 'Hallo, die Datei S:\Developers\b-end\\' + currentdate + '.xlsx ist fertig. \\r\\n Viele Grüße \\r\\n'
mail.Display(True)
